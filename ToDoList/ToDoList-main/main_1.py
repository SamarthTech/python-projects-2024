import time
from tkinter import *
from PIL import ImageTk,Image
from tkinter import font
import openpyxl 
from openpyxl import load_workbook
from tkinter import messagebox
from tkinter import ttk
import pandas as pd
class Window(Tk):
    def __init__(self):
        super().__init__()
        report_path = 'Assets\\user_data.xlsx'
        self.user_data = load_workbook(report_path)
        self.configure(cursor="cross")
        self.wm_iconbitmap(r"Assets\icon.ico")
        self.sheet_1 = self.user_data["Sheet1"]
        self.active_user = dict()
        self.title("To-do-list")
        self.geometry("800x500")
        self.font_all = font.Font(family="Comic Sans MS",weight="bold")
        self.tree_font = font.Font(family="Century Gothic",weight="normal",size=10)
        self.bg_image1 = ImageTk.PhotoImage(Image.open(r"Assets\bg_1.jpg"))
        self.resizable(0,0) 
        self.label_1 = Label(self, image=self.bg_image1)
        self.label_1.place(x=0, y=0)
        self.label_2 = Label(self, text="Hello User ,Please enter your name and password down below to proceed",
                             font=self.font_all, bg="black", fg="white")
        self.label_2.place(x=20, y=30)
        self.new_lb_1 = Label(self, text="Name", font=self.font_all, bg="black", fg="white")
        self.new_lb_2 = Label(self, text="Password", font=self.font_all, bg="black", fg="white")
        self.new_lb_1.place(x=60, y=150)
        self.entry_1 = Entry(self, width=30, bd=10, fg="black", bg="orange", font=self.font_all)
        self.entry_1.place(x=250, y=150)
        self.button_1 = Button(self, text="PROCEED", bg="lime green", fg="black", font=self.font_all,
                               activebackground="black", activeforeground="white", bd=10, relief=RIDGE,
                               command=lambda: self.get_user())
        self.button_1.place(x=350, y=325)
        self.entry_2 = Entry(self, bd=10, fg="black", show='*', bg="orange", font=self.font_all)
        self.new_lb_2.place(x=60, y=240)
        self.entry_2.place(x=250, y=240)
        self.update()
    def get_user(self):
        global user_def_run
        global current_username
        global pass_code
        global user_at_row
        user_def_run = False
        current_username = self.entry_1.get()
        pass_code = self.entry_2.get()
        new_user_val = True
        for i in range(1,self.sheet_1.max_row+1):
            if (self.sheet_1.cell(column=1,row=i).value == current_username):
                user_at_row = i
                new_user_val = False
            else:
                new_user_val = True

        if (new_user_val):
            self.new_user()
            self.mainframe_1()
        else:
            self.existing_user()
            self.mainframe_1()

    def new_user(self):
        global next_frame_render
        self.sheet_1.append((current_username,"[]","[]","[]",pass_code))
        self.user_at = self.sheet_1.max_row+1
        self.user_data.save(r"Assets\user_data.xlsx")
        self.active_user[current_username] = [[""],[""],[""]]
        user_def_run = True
        next_frame_render = True
    def existing_user(self):
        global next_frame_render
        if (pass_code == self.sheet_1.cell(row=user_at_row,column=5).value):
            for i in range(2,self.sheet_1.max_row+1):
                if (self.sheet_1.cell(column=1,row=i).value == current_username):
                    existing_username = self.sheet_1.cell(column=1,row=i)
                    self.user_at = i
                    break
            user_attribute = []
            for j in range(2,self.sheet_1.max_column+1):
                 user_attribute.append(self.sheet_1.cell(column=j,row=self.user_at).value)
            self.active_user[current_username] = user_attribute
            next_frame_render = True
            user_def_run  =True
        else:
            messagebox.showerror("Invalid passcode","The entered username or password is incorrect ,make sure you have entered it correctly")
            next_frame_render = False
    def update_existing(self):
        for i in range(2, self.sheet_1.max_row + 1):
            if (self.sheet_1.cell(column=1, row=i).value == current_username):
                existing_username = self.sheet_1.cell(column=1, row=i)
                self.user_at = i
                break
        user_attribute = []
        for j in range(2, self.sheet_1.max_column + 1):
            user_attribute.append(self.sheet_1.cell(column=j, row=self.user_at).value)
        self.active_user[current_username] = user_attribute
    def mainframe_1(self):
        try:
            global next_frame_render
            global user_def_run
            if (next_frame_render):
                w=0
                h=0
                frame_0 = Frame(self,width=w,height=h,bg="#36393E")
                for i in range(10):
                    w+=70
                    h+=46
                    frame_0.config(width=w,height=h)
                    frame_0.place(x=0,y=0)
                    time.sleep(0.1)
                    self.update()
                frame_1 = Frame(self,width=800,height=500)
                label_3 = Label(frame_1,image=self.bg_image1)
                label_3.place(x=0,y=0)
                stat_1 = "Hello "+str(current_username).capitalize()+"\n,You can review your tasks and make respective changes here"
                label_4 = Label(frame_1,text=stat_1,bg="black",fg="white",font=self.font_all)
                label_4.place(x=20,y=30)
                frame_1.place(x=0,y=0)
                #backend treeview
                works = []
                status_of_work = []
                time_ref = []
                treeview_1_values = dict()
                works_range = 0
                for i in (self.active_user.values()):
                    for j in eval(i[0]):
                        works.append(j)
                        works_range+=1
                for i in (self.active_user.values()):
                    for j in eval(i[1]):
                        status_of_work.append(j)
                for i in (self.active_user.values()):
                    for j in eval(i[2]):
                        time_ref.append(j)
                for i in range(works_range):
                    treeview_1_values[i] = (works[i],status_of_work[i],time_ref[i])
                # Scrollbar for treeview
                scrollbar_1 = ttk.Scrollbar(frame_1, orient="vertical")
                scrollbar_1.place(x=484,y=200,height=229)
                #treeview configurations
                treeview_1_header = ("Tasks","Status","Time Interval")
                treeview_1_config = ttk.Style()
                treeview_1_config.theme_use("clam")
                treeview_1_config_h = ttk.Style()
                treeview_1_config_h.configure("Treeview.heading")
                treeview_1_config.configure("Treeview", background="#36393E", foreground="silver",fieldbackground="#36393E",font=self.tree_font,yscrollcommand=scrollbar_1.set)
                treeview_1_config.map("Treeview",background=[("selected","#808080")],foreground=[("!selected","white"),("selected","black")])
                treeview_1_config.map("Treeview.heading")
                treeview_1 = ttk.Treeview(frame_1,show="headings")
                treeview_1["columns"] = treeview_1_header
                treeview_1.column("#0",width=1,anchor=W)
                treeview_1.column("Tasks",width=120,anchor=W)
                treeview_1.column("Status", width=120,anchor=W)
                treeview_1.column("Time Interval", width=120,anchor=W)
                treeview_1.heading("#0",text="",anchor=W)
                treeview_1.heading("Tasks", text="Tasks", anchor=W)
                treeview_1.heading("Status", text="Status", anchor=W)
                treeview_1.heading("Time Interval", text="Time Interval", anchor=W)
                button_1 = Button(frame_1,text="ADD",font=self.font_all,fg="black",bg="lime green",bd=5,relief=RIDGE,activebackground="black",activeforeground="white",width=7,command=lambda:add())
                button_1.place(x=584,y=220)
                button_2 = Button(frame_1,text="REMOVE",font=self.font_all,fg="black",bg="lime green",bd=5,relief=RIDGE,activebackground="black",activeforeground="white",width=7,command=lambda:remove())
                button_2.place(x=584,y=278)
                button_3 = Button(frame_1,text="MODIFY",font=self.font_all,fg="black",bg="lime green",bd=5,relief=RIDGE,activebackground="black",activeforeground="white",width=7,command=lambda:modify())
                button_3.place(x=584,y=336)
                scrollbar_1.configure(command=treeview_1.yview)
                count_1=0
                for i in treeview_1_values.values():
                    treeview_1.insert("",END,values=i,iid=count_1)
                    count_1+=1
                treeview_1.place(x=120,y=200)
                self.update()
                mainframe_1.update()
        except TypeError:
            self.destroy()
            self.__init__()
            messagebox.showinfo("Saved", "Greetings user ,your name has been successfully created")






        def remove():
            try:
                remove_item = treeview_1.focus()
                del works[int(remove_item)]
                del status_of_work[int(remove_item)]
                del time_ref[int(remove_item)]
                self.sheet_1.cell(row=self.user_at, column=2).value = str(works)
                self.sheet_1.cell(row=self.user_at, column=3).value = str(status_of_work)
                self.sheet_1.cell(row=self.user_at, column=4).value = str(time_ref)
                self.user_data.save(r"Assets\user_data.xlsx")
                self.update_existing()
                self.mainframe_1()
                mainframe_1.update()
            except ValueError:
                messagebox.showerror("ValueError","Please select an element from the tabulation")

        def modify_be():
            modified_work = str(entry_4.get())
            modified_time = str(modify_spinbox_hours_initial.get())+":"+str(modify_spinbox_minutes_initial.get())+":"+"00"+"-"+str(modify_spinbox_hours_final.get())+":"+str(modify_spinbox_minutes_final.get())+":"+"00"
            rework_addtime = modified_time.split("-")
            initial_rework_addtime = rework_addtime[0].split(":")
            final_rework_addtime = rework_addtime[1].split(":")
            for i in range(len(initial_rework_addtime)-1):
                if len(initial_rework_addtime[i]) == 1:
                    initial_rework_addtime[i] = "0"+initial_rework_addtime[i]
                    initial_rework_addtime[i]+=":"
                else:
                    initial_rework_addtime[i]+=":"

            for j in range(len(final_rework_addtime)-1):
                if len(final_rework_addtime[j])== 1:
                    final_rework_addtime[j] = "0"+final_rework_addtime[j]
                    final_rework_addtime[j]+=":"
                else:
                    final_rework_addtime[j] += ":"
            reworked_add_time = ""
            for i in initial_rework_addtime:
                reworked_add_time+=i
            reworked_add_time+="-"
            for j in final_rework_addtime:
                reworked_add_time+=j
            modify_status_of_work = modify_status_of_work_update.get()
            if modify_status_of_work == 1:
                modify_status_of_work = "done"
            else:
                modify_status_of_work = "not done"
            works[int(modify_item)] = modified_work
            time_ref[int(modify_item)] = reworked_add_time
            status_of_work[int(modify_item)] = modify_status_of_work
            self.sheet_1.cell(row=self.user_at,column=2).value = str(works)
            self.sheet_1.cell(row=self.user_at,column=3).value = str(status_of_work)
            self.sheet_1.cell(row=self.user_at,column=4).value = str(time_ref)
            self.user_data.save(r"Assets\user_data.xlsx")
            self.update_existing()
            self.mainframe_1()

        def modify():
            global modify_item
            global mainframe_3
            global entry_4
            global modify_spinbox_hours_initial
            global modify_spinbox_minutes_initial
            global modify_spinbox_hours_final
            global modify_spinbox_minutes_final
            global checkbutton_2
            global button_5
            global modify_status_of_work_update
            modify_item = treeview_1.focus()
            try:
                mainframe_3 = Toplevel(mainframe_1)
                mainframe_3.lift()
                mainframe_3.attributes("-topmost", True)
                mainframe_3.geometry("500x240")
                mainframe_3.iconbitmap(r"Assets\icon.ico")
                mainframe_3.resizable(0, 0)
                mainframe_3.config(bg="black")
                entry_4 = Entry(mainframe_3, bg="lime green", fg="black", font=self.font_all, width=25, bd=5, relief=RIDGE)
                entry_4.insert(END, str(works[int(modify_item)]))
                entry_4.place(x=85, y=20)
                modify_spinbox_hours_initial = Spinbox(mainframe_3, from_=0, to=24, bg="lime green", fg="black",
                                                font=self.font_all, bd=5, relief=RIDGE, width=8)
                modify_spinbox_minutes_initial = Spinbox(mainframe_3, from_=0, to=60, bg="lime green", fg="black",
                                                  font=self.font_all, bd=5, relief=RIDGE, width=8)
                modify_spinbox_hours_final = Spinbox(mainframe_3, from_=0, to=24, bg="lime green", fg="black", font=self.font_all,
                                              bd=5, relief=RIDGE, width=8)
                modify_spinbox_minutes_final = Spinbox(mainframe_3, from_=0, to=60, bg="lime green", fg="black",
                                                font=self.font_all, bd=5, relief=RIDGE, width=8)
                modify_spinbox_hours_initial.place(x=20, y=85)
                modify_spinbox_minutes_initial.place(x=155, y=85)
                modify_spinbox_hours_final.place(x=20, y=135)
                modify_spinbox_minutes_final.place(x=155, y=135)
                modify_status_of_work_update = IntVar()
                checkbutton_style = ttk.Style()
                checkbutton_style.configure("def.TCheckbutton", background="lime green", foreground="black",
                                            font=self.font_all)
                checkbutton_style.theme_use("clam")
                checkbutton_2 = ttk.Checkbutton(mainframe_3, text="Status", variable=modify_status_of_work_update, onvalue=1,
                                                offvalue=0, style="def.TCheckbutton")
                checkbutton_2.place(x=325, y=115)
                button_5 = Button(mainframe_3, text="MODIFY", fg="black", bg="lime green", relief=RIDGE, bd=5,
                                  activebackground="black", activeforeground="white", font=self.font_all, width=6,
                                  command=lambda: modify_be())
                button_5.place(x=395, y=175)
                mainframe_3.update()
            except(ValueError):
                messagebox.showerror(title="Invalid",message="Please select an option from the listbox")
                mainframe_3.destroy()

        def append_be():
            # task append
            add_task = entry_3.get()
            add_time = str(spinbox_hours_initial.get()) + ":" + str(
                spinbox_minutes_initial.get()) + ":" + "00" + "-" + str(spinbox_hours_final.get()) + ":" + str(
                spinbox_minutes_final.get()) + ":" + "00"
            rework_addtime = add_time.split("-")
            initial_rework_addtime = rework_addtime[0].split(":")
            final_rework_addtime = rework_addtime[1].split(":")
            for i in range(len(initial_rework_addtime) - 1):
                if len(initial_rework_addtime[i]) == 1:
                    initial_rework_addtime[i] = "0" + initial_rework_addtime[i]
                    initial_rework_addtime[i] += ":"
                else:
                    initial_rework_addtime[i] += ":"
            for j in range(len(final_rework_addtime) - 1):
                if len(final_rework_addtime[j]) == 1:
                    final_rework_addtime[j] = "0" + final_rework_addtime[j]
                    final_rework_addtime[j] += ":"
                else:
                    final_rework_addtime[j] += ":"
            reworked_add_time = ""
            for i in initial_rework_addtime:
                reworked_add_time += i
            reworked_add_time += "-"
            for j in final_rework_addtime:
                reworked_add_time += j
            add_status_of_work = status_of_work_update.get()
            if add_status_of_work == 1:
                add_status_of_work = "done"
            else:
                add_status_of_work = "not done"
            works.append(add_task)
            time_ref.append(reworked_add_time)
            status_of_work.append(add_status_of_work)
            self.sheet_1.cell(row=self.user_at, column=2).value = str(works)
            self.sheet_1.cell(row=self.user_at, column=3).value = str(status_of_work)
            self.sheet_1.cell(row=self.user_at, column=4).value = str(time_ref)
            self.user_data.save(r"Assets\user_data.xlsx")
            self.update_existing()
            self.mainframe_1()
        def add():
            global mainframe_2
            global entry_3
            global spinbox_hours_initial
            global spinbox_minutes_initial
            global spinbox_hours_final
            global spinbox_minutes_final
            global checkbutton_1
            global button_4
            global status_of_work_update
            mainframe_2 = Toplevel(mainframe_1)
            mainframe_2.lift()
            mainframe_2.attributes("-topmost",True)
            mainframe_2.geometry("500x240")
            mainframe_2.iconbitmap(r"Assets\icon.ico")
            mainframe_2.resizable(0, 0)
            mainframe_2.config(bg="black")
            entry_3 = Entry(mainframe_2,bg="lime green",fg="black",font=self.font_all,width=25,bd=5,relief=RIDGE)
            entry_3.insert(END,"Task Name")
            entry_3.place(x=85,y=20)
            spinbox_hours_initial = Spinbox(mainframe_2,from_=0,to=24,bg="lime green",fg="black",font=self.font_all,bd=5,relief=RIDGE,width=8)
            spinbox_minutes_initial = Spinbox(mainframe_2,from_=0,to=60,bg="lime green",fg="black",font=self.font_all,bd=5,relief=RIDGE,width=8)
            spinbox_hours_final = Spinbox(mainframe_2,from_=0,to=24,bg="lime green",fg="black",font=self.font_all,bd=5,relief=RIDGE,width=8)
            spinbox_minutes_final = Spinbox(mainframe_2,from_=0,to=60,bg="lime green",fg="black",font=self.font_all,bd=5,relief=RIDGE,width=8)
            spinbox_hours_initial.place(x=20,y=85)
            spinbox_minutes_initial.place(x=155,y=85)
            spinbox_hours_final.place(x=20,y=135)
            spinbox_minutes_final.place(x=155,y=135)
            status_of_work_update = IntVar()
            checkbutton_style = ttk.Style()
            checkbutton_style.configure("def.TCheckbutton",background="lime green",foreground="black",font=self.font_all,)
            checkbutton_style.theme_use("clam")
            checkbutton_1 = ttk.Checkbutton(mainframe_2,text="Status",variable=status_of_work_update,onvalue=1,offvalue=0,style="def.TCheckbutton")
            checkbutton_1.place(x=325,y=115)
            button_4 = Button(mainframe_2,text="ADD",fg="black",bg="lime green",relief=RIDGE,bd=5,activebackground="black",activeforeground="white",font=self.font_all,width=6,command=lambda :append_be())
            button_4.place(x=395,y=175)
            mainframe_2.update()


if __name__ == '__main__':
    mainframe_1 = Window()
    mainframe_1.mainloop()


